from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
import time
from selenium.webdriver.support.select import Select
import xlsxwriter
from openpyxl import Workbook
from openpyxl import load_workbook
import requests
from bs4 import BeautifulSoup
from datetime import datetime
import threading


stop_event = threading.Event()
start_time = time.time()

def get_input():
    while True:
        user_input = input("")
        print("The program will stop shortly.")
        if user_input == 'q' or user_input == 'quit':
            stop_event.set()  # Signal the main thread to stop
            return

# Create a thread to handle user input
input_thread = threading.Thread(target=get_input, daemon=True)
input_thread.start()


opt = Options()
opt.add_experimental_option("debuggerAddress", "localhost:9014")
browser = webdriver.Chrome(options=opt)


# school_year = 'P1_SCHOOL_YEAR'
report_level = 'P1_REPORT_LEVEL'
district = 'P1_DIVISION' # Division is district in virginia
school = 'P1_SCHOOL'
race = 'P1_RACE'
gender = 'P1_GENDER'
grade = 'P1_GRADE'
disadvantaged = 'P1_DISADVANTAGED'
english_learner = 'P1_LIMITED_ENGLISH'
migrant = 'P1_MIGRANT'
homeless = 'P1_HOMELESS'
military = 'P1_MILITARY'
foster_care = 'P1_FOSTER_CARE'
disabled = 'P1_DISABLED'

test_source = 'P1_TEST_SOURCE'
subject_area = 'P1_SUBJECT_AREA'
statistic = 'P1_STATISTIC'
test_test = 'P1_TEST'

district_checkpoint = 0  # For first run set to 1
school_checkpoint = 0  # For first run set to 1
grade_checkpoint = 0  # For first run set to 0
race_checkpoint = 0  # always 0
subject_area_checkpoint = 0  # For first run set to 0

gender_checkpoint = 0
disadvantaged_checkpoint = 0
english_learner_checkpoint = 0
former_english_learner_checkpoint = 0
migrant_checkpoint = 0
homeless_checkpoint = 0
military_checkpoint = 0
foster_care_checkpoint = 0
disabled_checkpoint = 0

file_path_name = FILE_PATH # change this to an existing xlsx file

workbook = load_workbook(file_path_name)
worksheet = workbook.active

time.sleep(3)

def check_loading():
    time.sleep(.5)
    while True:
        try:
            browser.find_element(By.CLASS_NAME, 'u-Processing')
        except:
            break

def selector(type, num, deselect=True):
    my_select = Select(browser.find_element(By.ID, type))
    if deselect:
        my_select.deselect_all()

    my_select.select_by_index(num)

def table_exists():
    page_source = browser.page_source
    soup = BeautifulSoup(page_source, 'html.parser')
    return not ('There is no data for this report.' in soup.text)


def verification():
    WebDriverWait(browser, 1000).until(EC.presence_of_element_located((By.ID, 'vdoePublicReCapchaFormSubmit')))
    browser.find_element(By.ID, 'vdoePublicReCapchaFormSubmit').click()
    WebDriverWait(browser, 1000).until(EC.presence_of_element_located((By.ID, 'P1_RESET')))

def set_constants():
    global grade_checkpoint

    browser.find_element(By.ID, 'P1_RESET').click()
    WebDriverWait(browser, 1000).until(EC.presence_of_element_located((By.ID, 'P1_RESET')))
    time.sleep(1)

    browser.execute_script(
        "document.getElementById('t_Header').style.display = 'none'")  # The header covers elements sometimes so I'm hiding it with js
    time.sleep(.5)

    # For schools:
    selector(report_level, 2)

    selector(test_source, 1)
    time.sleep(1)

    num_statistics = len(browser.find_element(By.ID, statistic).find_elements(By.TAG_NAME, 'option'))

    for statistic_num in range(0, num_statistics):
        selector(statistic, statistic_num, False)

    check_loading()

    selector(subject_area, subject_area_checkpoint)
    print("subject_area_checkpoint 1")
    print(subject_area_checkpoint)

    for grade_number in range(3, 9):
        selector(grade, grade_number, grade_number == 3)

    get_checkpoints(True)

def get_checkpoints(final_checkpoint):
    elapsed_time = time.time() - start_time

    hours = int(elapsed_time // 3600)
    minutes = int((elapsed_time % 3600) // 60)
    seconds = int(elapsed_time % 60)

    print('--------------------------------------')
    if final_checkpoint:
        print('CHECKPOINTS FINAL')
    else:
        print('CHECKPOINTS')
    print("district_checkpoint: " + str(district_checkpoint))
    print("school_checkpoint: " + str(school_checkpoint))
    print("grade_checkpoint: " + str(grade_checkpoint))
    print("subject_area_checkpoint: " + str(subject_area_checkpoint))
    print(f'Program has been running for: {hours:02}:{minutes:02}:{seconds:02}')
    print('--------------------------------------')


def collect_to_xlsx(data, file_path=file_path_name):
    if worksheet.max_row == 0 or not any(cell.value for cell in worksheet[1]):
        excel_first_row = []
    else:
        excel_first_row = [cell.value for cell in worksheet[1]]

    first_row_data = data[0]
    column_numbers = []

    for value in first_row_data:
        if value in excel_first_row:
            column_numbers.append(excel_first_row.index(value) + 1)
        else:
            next_empty_col = len(excel_first_row) + 1
            worksheet.cell(row=1, column=next_empty_col, value=value)
            excel_first_row.append(value)
            column_numbers.append(next_empty_col)

    print("Column numbers for the first row of data:", column_numbers)

    for row_data in data[1:]:
        row = worksheet.max_row + 1
        for i, value in enumerate(row_data):
            column = column_numbers[i]
            worksheet.cell(row=row, column=column, value=value)
    workbook.save(file_path)

def get_invis_cols(rest_of_rows, type, name, checkpoint_name):
    # Ex. of name is "Grade" (something written in the table)
    if name not in rest_of_rows[0]:
        rest_of_rows[0].append(name)
        for row_id in range(len(rest_of_rows)):
            if row_id == 0:
                continue
            rest_of_rows[row_id].append(checkpoint_name)
    return rest_of_rows

def get_table():
    if not table_exists():
        print('NOTHING EXISTS FOR THIS')
        return [[]]
    first_row = [i.text for i in browser.find_element(By.TAG_NAME, 'thead').find_element(By.TAG_NAME, 'tr').find_elements(By.TAG_NAME, 'th')]
    rows = browser.find_element(By.CLASS_NAME, 't-Report-report').find_element(By.TAG_NAME, 'tbody').find_elements(By.TAG_NAME, 'tr')
    rest_of_rows = [first_row]
    for row in rows:
        rest_of_rows.append([item.text for item in row.find_elements(By.TAG_NAME, 'td')])

    rest_of_rows = get_invis_cols(rest_of_rows, grade, 'Grade', "All Grades")
    rest_of_rows = get_invis_cols(rest_of_rows, race, 'Race', "All Races")

    print(rest_of_rows)
    return rest_of_rows


def submit_values():
    WebDriverWait(browser, 1000).until(EC.presence_of_element_located((By.ID, 'P1_SUBMIT')))
    # print('loaded')
    browser.find_element(By.ID, 'P1_SUBMIT').click()
    time.sleep(.5)
    WebDriverWait(browser, 1000).until(EC.presence_of_element_located((By.ID, 'P1_SUBMIT')))
    time.sleep(.5)

    collect_to_xlsx(get_table())
    multiple_pages()

    for grade_number in range(3, 9):
        selector(grade, grade_number, grade_number == 3)


def multiple_pages():
    num_pages = 0
    while True:
        try:
            browser.find_element(By.CLASS_NAME, 't-Report-paginationLink--next').click()
            check_loading()
            collect_to_xlsx(get_table())
            num_pages += 1
        except:
            break
    if num_pages > 1:
        reset()

def reset():
    browser.refresh()
    browser.get('https://p1pe.doe.virginia.gov/apex_captcha/home.do?apexTypeId=306')
    time.sleep(.5)
    WebDriverWait(browser, 1000).until(EC.presence_of_element_located((By.ID, 'vdoePublicReCapchaFormSubmit')))
    time.sleep(1)
    verification()
    set_constants()


def all_sub_sections_runthrough():
    global subject_area_checkpoint, race_checkpoint, gender_checkpoint, disadvantaged_checkpoint, english_learner_checkpoint, former_english_learner_checkpoint
    global migrant_checkpoint, homeless_checkpoint, military_checkpoint, foster_care_checkpoint, disabled_checkpoint
    num_subject_areas = 5

    for subject_area_num in range(subject_area_checkpoint, num_subject_areas):
        if subject_area_num == 2:
            subject_area_checkpoint += 1
            continue
        time.sleep(.5)
        selector(subject_area, subject_area_num)

        check_loading()

        time.sleep(1)
        if subject_area_num == 0:
            selector(test_test, 1)
        if subject_area_num == 1:
            selector(test_test, 1)
        if subject_area_num == 3:
            selector(test_test, 4)
        if subject_area_num == 4:
            selector(test_test, 4)


        if disadvantaged_checkpoint == 0:
            selector(disadvantaged, 2)
            submit_values()
            selector(disadvantaged, 0)
            disadvantaged_checkpoint = 2

        if english_learner_checkpoint == 0:
            selector(english_learner, 2)
            submit_values()
            selector(english_learner, 0)
            english_learner_checkpoint = 2

        if migrant_checkpoint == 0:
            selector(migrant, 2)
            submit_values()
            selector(migrant, 0)
            migrant_checkpoint = 2

        if homeless_checkpoint == 0:
            selector(homeless, 2)
            submit_values()
            selector(homeless, 0)
            homeless_checkpoint = 2

        if military_checkpoint == 0:
            selector(military, 2)
            submit_values()
            selector(military, 0)
            military_checkpoint = 2

        if foster_care_checkpoint == 0:
            selector(foster_care, 2)
            submit_values()
            selector(foster_care, 0)
            foster_care_checkpoint = 2

        if disabled_checkpoint == 0:
            selector(disabled, 2)
            submit_values()
            selector(disabled, 0)
            disabled_checkpoint = 2

        if subject_area_checkpoint != 4:
            subject_area_checkpoint += 1

        gender_checkpoint = 0
        disadvantaged_checkpoint = 0
        english_learner_checkpoint = 0
        former_english_learner_checkpoint = 0
        migrant_checkpoint = 0
        homeless_checkpoint = 0
        military_checkpoint = 0
        foster_care_checkpoint = 0
        disabled_checkpoint = 0

        get_checkpoints(False)

        if stop_event.is_set():
            quit()


browser.refresh()
browser.get('https://p1pe.doe.virginia.gov/apex_captcha/home.do?apexTypeId=306')
verification()
set_constants()


num_grades = len(browser.find_element(By.ID, grade).find_elements(By.TAG_NAME, 'option'))


# selector(grade, 0)

for grade_number in range(3, 9):
    selector(grade, grade_number, grade_number == 3)

all_sub_sections_runthrough()
